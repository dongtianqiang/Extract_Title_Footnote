"""RTF Content Processor - Integrated Workflow.

This script provides an integrated workflow that:
1. Reads a LOT file (xlsx format) and extracts filenames from '文件名称' column
2. Creates a temp_file folder in the LOT file's directory
3. Copies RTF files listed in the LOT file to the temp_file folder
4. Processes each RTF file by removing content between specific markers:
   - From the line AFTER the first line containing "\\cell}"
   - To the line BEFORE the last line containing "\\pard\\plain\\qc"
   (Boundary lines are NOT removed)
5. Merges all processed RTF files into a single file according to rules:
   - For all RTF files except the last one: remove the last "}" character
   - For all RTF files except the first one: keep only the line containing "{\\header\\pard" and content below
   - Merge all contents without adding any extra information

Usage:
    python process_rtf_content.py                           # Uses default paths
    python process_rtf_content.py <lot_path> <output_path>  # Uses specified paths

The script uses default paths:
    - LOT file: .\\example\\rtf\\test_LOT_file.xlsx
    - Output: .\\example\\rtf\\title_footnote_rtf.rtf
    
Or when using command line arguments:
    - lot_path: Path to the LOT file
    - output_path: Path for the output merged RTF file
"""

import os
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def find_rtf_files(directory):
    """
    Find all RTF files in the specified directory.
    
    Args:
        directory: Path to search for RTF files
        
    Returns:
        List of paths to RTF files
    """
    dir_path = Path(directory)
    rtf_files = list(dir_path.glob("*.rtf"))
    return rtf_files


def process_rtf_file(rtf_path):
    """
    Process a single RTF file by removing content between markers.
    
    Removes content from the line AFTER the first occurrence of a line
    containing "\\cell}" to the line BEFORE the last occurrence of a line
    containing "\\pard\\plain\\qc" (boundary lines are preserved).
    
    Args:
        rtf_path: Path to the RTF file
        
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        # Read the RTF file
        with open(rtf_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        if not lines:
            return False, "File is empty"
        
        # Find the first line containing "\cell}"
        first_cell_line = None
        for idx, line in enumerate(lines):
            if '\\cell}' in line:
                first_cell_line = idx
                break
        
        # Find the last line containing "\pard\plain\qc"
        last_pard_qc_line = None
        for idx in range(len(lines) - 1, -1, -1):
            if '\\pard\\plain\\qc' in lines[idx]:
                last_pard_qc_line = idx
                break
        
        # Validate that both markers were found
        if first_cell_line is None:
            return False, "No line with '\\cell}' found"
        
        if last_pard_qc_line is None:
            return False, "No line with '\\pard\\plain\\qc' found"
        
        # Check if the range is valid (need at least one line to remove)
        # Start removal from first_cell_line + 1 (after \cell} line)
        # End removal at last_pard_qc_line - 1 (before \pard\plain\qc line)
        start_removal = first_cell_line + 1
        end_removal = last_pard_qc_line - 1
        
        if start_removal > end_removal:
            print(f"Invalid range: start at line {first_cell_line + 1}, end at line {last_pard_qc_line + 1}")
                
        # Create new content by keeping:
        # 1. Everything up to and including first_cell_line
        # 2. Everything from last_pard_qc_line onwards
        new_lines = lines[:start_removal] + lines[end_removal + 1:]
        
        # Find all lines containing "\trowd\trkeep\trql" in the new content
        # We need to find indices of '\trowd\trkeep\trql' that appear AFTER the last '\pard\plain\qc' line.
        # The last '\pard\plain\qc' line is located at index `first_cell_line + 1` in `new_lines`.
        
        last_pard_qc_index_in_new = first_cell_line + 1
        
        # Find all indices of '\trowd\trkeep\trql' strictly after the last '\pard\plain\qc' line
        target_marker = '\\trowd\\trkeep\\trql'
        marker_indices = [i for i, line in enumerate(new_lines) if i > last_pard_qc_index_in_new and target_marker in line]
        
        intbl_message = ""
        # If we have at least 2 occurrences, remove from the second-to-last occurrence to the second-to-last line of the file (inclusive)
        if len(marker_indices) >= 2:
            start_remove_idx = marker_indices[-2]
            # The end boundary is the second-to-last line of the current new_lines list
            end_remove_idx = len(new_lines) - 2
            
            # Ensure the range is valid (start <= end)
            if start_remove_idx <= end_remove_idx:
                # Delete the range [start_remove_idx, end_remove_idx] inclusive
                count_to_remove = end_remove_idx - start_remove_idx + 1
                del new_lines[start_remove_idx : end_remove_idx + 1]
                
                #intbl_message = f", removed {count_to_remove} lines from 2nd-to-last '{target_marker}' to the second-to-last line (inclusive)"
            #else:
                #intbl_message = f", skipped removal as 2nd-to-last '{target_marker}' is after the second-to-last line"

        # Write the modified content back
        with open(rtf_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
                
        removed_count = (end_removal + 1) - start_removal
        return True, f"Removed {removed_count} lines (from line {start_removal + 1} to {end_removal + 1}){intbl_message}"
        
    except Exception as e:
        return False, f"Error processing file: {str(e)}"


def process_directory(directory):
    """
    Process all RTF files in a directory.
    
    Args:
        directory: Path to directory containing RTF files
        
    Returns:
        Dictionary with processing results
    """
    dir_path = Path(directory)
    
    if not dir_path.exists():
        raise FileNotFoundError(f"Directory not found: {directory}")
    
    # Find all RTF files
    rtf_files = find_rtf_files(directory)
    
    if not rtf_files:
        print(f"No RTF files found in {directory}")
        return {
            'total': 0,
            'success': 0,
            'failed': 0,
            'results': []
        }
    
    print(f"Found {len(rtf_files)} RTF file(s) to process")
    print("-" * 60)
    
    results = {
        'total': len(rtf_files),
        'success': 0,
        'failed': 0,
        'results': []
    }
    
    for rtf_file in rtf_files:
        print(f"\nProcessing: {rtf_file.name}")
        
        # Process the file (backup disabled by default)
        success, message = process_rtf_file(rtf_file)
        
        if success:
            print(f"{message}")
            results['success'] += 1
        else:
            print(f"{message}")
            results['failed'] += 1
        
        results['results'].append({
            'file': rtf_file.name,
            'success': success,
            'message': message
        })
    
    print("\n" + "=" * 60)
    print("Processing Complete!")
    print(f"Total files: {results['total']}")
    print(f"Successfully processed: {results['success']}")
    print(f"Failed: {results['failed']}")
    print("=" * 60)
    
    return results


def read_lot_file_for_merge(lot_path):
    """
    Read the LOT file and extract filenames from the '文件名称' column.
    
    Args:
        lot_path: Path to the LOT file (xlsx format)
        
    Returns:
        List of filenames from the '文件名称' column
    """
    lot_path = Path(lot_path)
    
    if not lot_path.exists():
        raise FileNotFoundError(f"LOT file not found: {lot_path}")
    
    print(f"\nReading LOT file: {lot_path}")
    
    # Load the workbook
    wb = load_workbook(filename=lot_path, read_only=True)
    ws = wb.active
    
    # Get the header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value if cell.value else "")
    
    print(f"Headers found: {headers}")
    
    # Find the '文件名称' column
    target_column = None
    for idx, header in enumerate(headers):
        if header == '文件名称':
            target_column = idx + 1  # 1-based index
            break
    
    if target_column is None:
        raise ValueError("Column '文件名称' not found in the LOT file")
    
    print(f"Found '文件名称' column at position: {target_column}")
    
    # Extract filenames from the '文件名称' column
    filenames = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=target_column, max_col=target_column), start=2):
        cell_value = row[0].value
        if cell_value:
            # Clean the filename - ensure it has .rtf extension
            filename = str(cell_value).strip()
            if not filename.lower().endswith('.rtf'):
                filename += '.rtf'
            filenames.append(filename)
    
    wb.close()
    
    print(f"Found {len(filenames)} filenames in LOT file")
    return filenames


def merge_rtf_files(rtf_file_list, output_path):
    """
    Merge RTF files according to specific rules:
    1. For all RTF files except the last one: remove the last "}" character
    2. For all RTF files except the first one: keep only the line containing "{\\header\\pard" and content below
    3. Merge all contents without adding any extra information
    
    Args:
        rtf_file_list: List of RTF file paths in order
        output_path: Path for the merged output file
    """
    if not rtf_file_list:
        print("No RTF files to merge")
        return False
    
    print(f"\nMerging {len(rtf_file_list)} RTF files...")
    merged_content = []
    
    for idx, rtf_file in enumerate(rtf_file_list):
        print(f"  Processing: {rtf_file.name} ({idx + 1}/{len(rtf_file_list)})")
        
        try:
            with open(rtf_file, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            if not lines:
                print(f"    Warning: File is empty, skipping...")
                continue
            
            # Rule 1: For all files except the last one, remove the last "}"
            if idx < len(rtf_file_list) - 1:
                # Find and remove only the very last "}" character in the file content.
                # We iterate backwards through lines to find the last line containing "}",
                # then remove only the rightmost "}" from that specific line.
                for line_idx in range(len(lines) - 1, -1, -1):
                    if '}' in lines[line_idx]:
                        # Strip trailing whitespace/newlines to isolate the character
                        stripped_line = lines[line_idx].rstrip('\r\n')
                        if stripped_line.endswith('}'):
                            # Remove only the last character '}'
                            new_line_content = stripped_line[:-1]
                            # Re-attach the original line ending (defaulting to \n if mixed/missing)
                            original_ending = lines[line_idx][len(stripped_line):]
                            if not original_ending:
                                original_ending = '\n'
                            lines[line_idx] = new_line_content + original_ending
                        break
            
            # Rule 2: For all files except the first one, keep only {\header\pard line and below
            if idx > 0:
                header_pard_line_idx = None
                for line_idx, line in enumerate(lines):
                    if '{\\header\\pard' in line:
                        header_pard_line_idx = line_idx
                        break
                
                if header_pard_line_idx is not None:
                    # Keep only from header_pard_line onwards
                    lines = lines[header_pard_line_idx:]
                    #print(f"    Kept content from {{\\header\\pard line onwards")
                else:
                    print(f"    Warning: No header line found, keeping all content")
            
            # Add the processed content to merged_content
            merged_content.extend(lines)
            
        except Exception as e:
            print(f"    Error processing file: {str(e)}")
            continue
    
    # Write the merged content
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.writelines(merged_content)
        print(f"\n Merged RTF file created: {output_path}")
        return True
    except Exception as e:
        print(f"\n Error writing merged file: {str(e)}")
        return False


def process_lot_and_merge_rtf(lot_path, output_file, cancellation_token=None):
    """
    Process LOT file and merge RTF files with cancellation support.
    
    This is a wrapper function that can be called from GUI with cancellation token.
    
    Args:
        lot_path: Path to the LOT file
        output_file: Path for the merged output RTF file
        cancellation_token: Token to check for cancellation requests
    
    Returns:
        bool: True if successful, False otherwise
    """
    import shutil
    
    try:
        lot_path = Path(lot_path)
        lot_dir = lot_path.parent
        temp_dir = lot_dir / "_tf_temp_folder"
        
        # Step 1: Read the LOT file
        print("\n" + "=" * 60)
        print("Step 1: Reading LOT file")
        print("=" * 60)
        filenames = read_lot_file_for_merge(lot_path)
        
        if not filenames:
            print("\nError: No filenames found in LOT file!")
            return False
        
        # Check cancellation
        if cancellation_token and cancellation_token.is_cancelled():
            print("\nOperation cancelled by user.")
            return False
        
        # Step 2: Check/Create temp_file folder
        print("\n" + "=" * 60)
        print("Step 2: Checking/Creating temp_file folder")
        print("=" * 60)
        if temp_dir.exists():
            print(f"temp_file folder already exists: {temp_dir}")
        else:
            print(f"\nCreating temp_file folder: {temp_dir}")
            temp_dir.mkdir(parents=True, exist_ok=True)
            print(f"temp_file folder created successfully")
        
        # Check cancellation
        if cancellation_token and cancellation_token.is_cancelled():
            print("\nOperation cancelled by user.")
            return False
        
        # Step 3: Copy RTF files to temp_file
        print("\n" + "=" * 60)
        print("Step 3: Copying RTF files to temp_file")
        print("=" * 60)
        print(f"Source: {lot_dir}")
        print(f"Destination: {temp_dir}")
        print("-" * 60)
        
        copied_count = 0
        missing_files = []
        for filename in filenames:
            # Check cancellation before each file operation
            if cancellation_token and cancellation_token.is_cancelled():
                print("\nOperation cancelled by user during file copy.")
                return False
            
            source_file = lot_dir / filename
            dest_file = temp_dir / filename
            
            if source_file.exists():
                shutil.copy2(source_file, dest_file)
                print(f"Copied: {filename}")
                copied_count += 1
            else:
                print(f"Missing: {filename}")
                missing_files.append(filename)
        
        print("-" * 60)
        print(f"\nCopied {copied_count} out of {len(filenames)} files")
        if missing_files:
            print(f"\nWarning: {len(missing_files)} file(s) not found (will skip these)")
        
        # Build the list of RTF files to process
        rtf_files = []
        for filename in filenames:
            rtf_path = temp_dir / filename
            if rtf_path.exists():
                rtf_files.append(rtf_path)
        
        if not rtf_files:
            print("\nError: No RTF files found to process!")
            return False
        
        # Check cancellation
        if cancellation_token and cancellation_token.is_cancelled():
            print("\nOperation cancelled by user.")
            return False
        
        # Step 4: Process each RTF file (remove content between markers)
        print("\n" + "=" * 60)
        print("Step 4: Processing RTF files (removing sections)")
        print("=" * 60)
        
        for rtf_file in rtf_files:
            # Check cancellation before processing each file
            if cancellation_token and cancellation_token.is_cancelled():
                print("\nOperation cancelled by user during RTF processing.")
                return False
            
            print(f"\nProcessing: {rtf_file.name}")
            success, message = process_rtf_file(rtf_file)
            if success:
                print(f"{message}")
            else:
                print(f"{message}")
        
        print("\n" + "=" * 60)
        print("RTF Content Processing Complete")
        print("=" * 60)
        
        # Check cancellation
        if cancellation_token and cancellation_token.is_cancelled():
            print("\nOperation cancelled by user.")
            return False
        
        # Step 5: Merge RTF files according to specified rules
        print("\n" + "=" * 60)
        print("Step 5: Merging RTF files")
        print("=" * 60)
        
        success = merge_rtf_files(rtf_files, output_file)
        
        if success:
            print("\n" + "=" * 60)
            print("Workflow Complete!")
            print(f"Output file: {output_file}")
            print("=" * 60)
            
            # Clean up temp_file folder
            print("\n" + "=" * 60)
            print("Step 6: Cleaning up temporary files")
            print("=" * 60)
            if temp_dir.exists():
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                    print(f"Temporary folder deleted: {temp_dir}")
                except Exception as cleanup_error:
                    print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
                    print(f"You can manually delete: {temp_dir}")
            else:
                print(f"Temporary folder not found (already deleted?): {temp_dir}")
            
            return True
        else:
            print("\nMerge failed!")
            return False
        
    except KeyboardInterrupt:
        # Handle cancellation gracefully
        print("\nOperation cancelled by user.")
        # Clean up temp_file folder even on cancellation
        if 'temp_dir' in locals() and temp_dir.exists():
            try:
                import shutil
                shutil.rmtree(temp_dir)
                print(f"Temporary folder deleted: {temp_dir}")
            except Exception as cleanup_error:
                print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
        return False
    except Exception as e:
        print(f"\nFatal error: {str(e)}")
        import traceback
        traceback.print_exc()
        # Clean up temp_file folder on error
        if 'temp_dir' in locals() and temp_dir.exists():
            try:
                import shutil
                shutil.rmtree(temp_dir)
                print(f"Temporary folder deleted due to error: {temp_dir}")
            except Exception as cleanup_error:
                print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
        return False


def main():
    """Main entry point - Integrated LOT and RTF processing workflow."""
    import sys
    import shutil
    
    print("=" * 60)
    print("RTF Content Processor - Integrated Workflow")
    print("=" * 60)
    
    # Check if command line arguments are provided
    if len(sys.argv) == 3:
        # Use command line arguments
        lot_file_path = sys.argv[1]
        output_file = sys.argv[2]
    else:
        print("Error: Expected 2 arguments (LOT file path and output file path)")
        print(f"Received {len(sys.argv) - 1} argument(s)")
        sys.exit(1)
    
    try:
        lot_path = Path(lot_file_path)
        lot_dir = lot_path.parent
        temp_dir = lot_dir / "_tf_temp_folder"
        
        # Step 1: Read the LOT file
        print("\n" + "=" * 60)
        print("Step 1: Reading LOT file")
        print("=" * 60)
        filenames = read_lot_file_for_merge(lot_file_path)
        
        if not filenames:
            print("\nError: No filenames found in LOT file!")
            sys.exit(1)
        
        # Step 2: Check/Create temp_file folder
        print("\n" + "=" * 60)
        print("Step 2: Checking/Creating temp_file folder")
        print("=" * 60)
        if temp_dir.exists():
            print(f"temp_file folder already exists: {temp_dir}")
        else:
            print(f"\nCreating temp_file folder: {temp_dir}")
            temp_dir.mkdir(parents=True, exist_ok=True)
            print(f"temp_file folder created successfully")
        
        # Step 3: Copy RTF files to temp_file
        print("\n" + "=" * 60)
        print("Step 3: Copying RTF files to temp_file")
        print("=" * 60)
        print(f"Source: {lot_dir}")
        print(f"Destination: {temp_dir}")
        print("-" * 60)
        
        copied_count = 0
        missing_files = []
        for filename in filenames:
            source_file = lot_dir / filename
            dest_file = temp_dir / filename
            
            if source_file.exists():
                shutil.copy2(source_file, dest_file)
                print(f"Copied: {filename}")
                copied_count += 1
            else:
                print(f"Missing: {filename}")
                missing_files.append(filename)
        
        print("-" * 60)
        print(f"\nCopied {copied_count} out of {len(filenames)} files")
        if missing_files:
            print(f"\nWarning: {len(missing_files)} file(s) not found (will skip these)")
        
        # Build the list of RTF files to process
        rtf_files = []
        for filename in filenames:
            rtf_path = temp_dir / filename
            if rtf_path.exists():
                rtf_files.append(rtf_path)
        
        if not rtf_files:
            print("\nError: No RTF files found to process!")
            sys.exit(1)
        
        # Step 4: Process each RTF file (remove content between markers)
        print("\n" + "=" * 60)
        print("Step 4: Processing RTF files (removing sections)")
        print("=" * 60)
        
        for rtf_file in rtf_files:
            print(f"\nProcessing: {rtf_file.name}")
            success, message = process_rtf_file(rtf_file)
            if success:
                print(f"{message}")
            else:
                print(f"{message}")
        
        print("\n" + "=" * 60)
        print("RTF Content Processing Complete")
        print("=" * 60)
        
        # Step 5: Merge RTF files according to specified rules
        print("\n" + "=" * 60)
        print("Step 5: Merging RTF files")
        print("=" * 60)
        
        success = merge_rtf_files(rtf_files, output_file)
        
        if success:
            print("\n" + "=" * 60)
            print("Workflow Complete!")
            print(f"Output file: {output_file}")
            print("=" * 60)
            
            # Clean up temp_file folder
            print("\n" + "=" * 60)
            print("Step 6: Cleaning up temporary files")
            print("=" * 60)
            if temp_dir.exists():
                try:
                    shutil.rmtree(temp_dir)
                    print(f"Temporary folder deleted: {temp_dir}")
                except Exception as cleanup_error:
                    print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
                    print(f"You can manually delete: {temp_dir}")
            else:
                print(f"Temporary folder not found (already deleted?): {temp_dir}")
            
            sys.exit(0)
        else:
            print("\nMerge failed!")
            sys.exit(1)
        
    except KeyboardInterrupt:
        # Handle cancellation gracefully
        print("\nOperation cancelled by user.")
        # Clean up temp_file folder even on cancellation
        if 'temp_dir' in locals() and temp_dir.exists():
            try:
                shutil.rmtree(temp_dir)
                print(f"Temporary folder deleted: {temp_dir}")
            except Exception as cleanup_error:
                print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        print(f"\nFatal error: {str(e)}")
        import traceback
        traceback.print_exc()
        # Clean up temp_file folder on error
        if 'temp_dir' in locals() and temp_dir.exists():
            try:
                shutil.rmtree(temp_dir)
                print(f"Temporary folder deleted due to error: {temp_dir}")
            except Exception as cleanup_error:
                print(f"Warning: Failed to delete temporary folder: {cleanup_error}")
        sys.exit(1)


if __name__ == "__main__":
    main()
